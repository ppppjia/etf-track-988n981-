import sys
import os
import requests
import io
import argparse

# 將目前目錄加入 Python 模組搜尋路徑
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from importer import parse_roc_date, import_excel_file

ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
DOWNLOADS_DIR = os.path.join(ROOT_DIR, "data", "downloaded")

URLS = {
    "00981A": "https://www.ezmoney.com.tw/ETF/Fund/AssetExcelNPOI?fundCode=49YTW",
    "00988A": "https://www.ezmoney.com.tw/ETF/Fund/AssetExcelNPOI?fundCode=61YTW"
}

def download_etf_portfolio(etf_code):
    """
    從網址下載投組 Excel，解析內容中的交易日期，以 '代號-交易日期.xlsx' 重命名儲存，並導入資料庫。
    """
    url = URLS.get(etf_code)
    if not url:
        raise ValueError(f"不支援的 ETF 代號: {etf_code}")
        
    print(f"正在從 {url} 下載 {etf_code} 投組...")
    
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    # 進行 HTTP 下載，停用 SSL 驗證以防憑證錯誤
    r = requests.get(url, headers=headers, verify=False, timeout=30)
    if r.status_code != 200:
        raise ConnectionError(f"下載失敗，HTTP 狀態碼: {r.status_code}")
        
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
    sheet = wb.active
    
    # 搜尋 "資料日期"
    date_val = None
    for r_idx in range(1, 4):
        for c_idx in range(1, 3):
            val = sheet.cell(row=r_idx, column=c_idx).value
            if val and "資料日期" in str(val):
                date_val = val
                break
        if date_val:
            break
            
    if not date_val:
        date_val = sheet.cell(row=1, column=1).value
        
    trading_date = parse_roc_date(date_val)
    if not trading_date:
        raise ValueError(f"無法從下載的 Excel 中解析出交易日期: {date_val}")
        
    date_str_formatted = trading_date.replace("-", "")
    short_code = etf_code.replace("00", "").replace("A", "") # '00981A' -> '981'
    
    new_filename = f"{short_code}-{date_str_formatted}.xlsx"
    
    # 存放在 data/downloaded 歸檔目錄
    os.makedirs(DOWNLOADS_DIR, exist_ok=True)
    archive_path = os.path.join(DOWNLOADS_DIR, new_filename)
    
    # 寫入檔案
    with open(archive_path, "wb") as f:
        f.write(r.content)
    print(f"已存檔至歸檔目錄: {archive_path}")
    
    # 執行資料庫匯入
    import_excel_file(archive_path, etf_code)
    
    return archive_path, trading_date

def run_downloader():
    """
    下載所有設定的 ETF 投組。
    """
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    results = {}
    for code in URLS:
        try:
            path, date = download_etf_portfolio(code)
            results[code] = {"status": "success", "path": path, "date": date}
        except Exception as e:
            print(f"下載/匯入 {code} 時發生錯誤: {e}")
            results[code] = {"status": "error", "error": str(e)}
            
    return results

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="主動型 ETF 投組每日下載器")
    parser.add_argument("--manual", action="store_true", help="手動觸發即時下載")
    args = parser.parse_args()
    
    print("啟動 ETF 投組下載程序...")
    res = run_downloader()
    print("下載程序結束。結果：", res)
