import openpyxl
import re
import os
from database import save_etf_data

def clean_number(val):
    """
    清理數字欄位，移除貨幣符號、千分位逗號、百分比符號，並轉為浮點數或整數。
    """
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return val
    
    val_str = str(val).strip()
    val_str = re.sub(r'[^\d\.\-]', '', val_str)
    
    if not val_str:
        return 0.0
    
    try:
        if '.' in val_str:
            return float(val_str)
        else:
            return int(val_str)
    except ValueError:
        return 0.0

def parse_roc_date(date_cell_val):
    """
    解析民國日期（例如 "資料日期：115/07/31" 或 "115/07/31"）並轉換為西元 "YYYY-MM-DD" 格式。
    """
    if not date_cell_val:
        return None
    
    match = re.search(r'(\d+)/(\d+)/(\d+)', str(date_cell_val))
    if not match:
        return None
    
    roc_year = int(match.group(1))
    month = int(match.group(2))
    day = int(match.group(3))
    
    ad_year = roc_year + 1911
    
    return f"{ad_year:04d}-{month:02d}-{day:02d}"

def import_excel_file(file_path, etf_code=None):
    """
    解析一個投組 Excel 檔案並匯入資料庫。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"檔案不存在: {file_path}")
        
    filename = os.path.basename(file_path)
    
    if etf_code is None:
        if "981" in filename:
            etf_code = "00981A"
        elif "988" in filename:
            etf_code = "00988A"
        else:
            raise ValueError(f"無法從檔名 {filename} 判斷 ETF 代號，請手動指定 etf_code。")
            
    print(f"正在解析 {filename} (代號: {etf_code})...")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    date_val = None
    for r in range(1, 4):
        for c in range(1, 3):
            val = sheet.cell(row=r, column=c).value
            if val and "資料日期" in str(val):
                date_val = val
                break
        if date_val:
            break
            
    if not date_val:
        date_val = sheet.cell(row=1, column=1).value
        
    trading_date = parse_roc_date(date_val)
    if not trading_date:
        raise ValueError(f"無法從檔案中解析出交易日期：{date_val}")
        
    print(f"解析到交易日期: {trading_date}")
    
    summary_data = {
        "net_assets": 0.0,
        "units_outstanding": 0.0,
        "nav": 0.0
    }
    
    for r in range(1, 20):
        cell_a = sheet.cell(row=r, column=1).value
        cell_b = sheet.cell(row=r, column=2).value
        
        if cell_a and isinstance(cell_a, str):
            cell_a_clean = cell_a.strip()
            if "淨資產" in cell_a_clean and "每單位" not in cell_a_clean:
                summary_data["net_assets"] = clean_number(cell_b)
            elif "流通在外單位數" in cell_a_clean or "單位數" in cell_a_clean:
                summary_data["units_outstanding"] = clean_number(cell_b)
            elif "每單位淨值" in cell_a_clean or "單位淨值" in cell_a_clean:
                summary_data["nav"] = clean_number(cell_b)
                
    print(f"解析到總結數據: {summary_data}")
    
    holdings = []
    start_row = None
    
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 5)]
        row_vals_str = [str(v).strip() if v is not None else "" for v in row_vals]
        
        if "股票代號" in row_vals_str or "代號" in row_vals_str:
            start_row = r + 1
            break
            
    if start_row is None:
        raise ValueError("找不到成分股表格標頭（應包含 '股票代號' 欄位）")
        
    for r in range(start_row, sheet.max_row + 1):
        code = sheet.cell(row=r, column=1).value
        name = sheet.cell(row=r, column=2).value
        shares = sheet.cell(row=r, column=3).value
        weight = sheet.cell(row=r, column=4).value
        
        if code is None or str(code).strip() == "" or str(code).strip().startswith("註"):
            break
            
        code_str = str(code).strip()
        name_str = str(name).strip() if name is not None else ""
        
        if code_str.lower() in ("tx", "合計", "項目", "現金"):
            continue
            
        clean_shares = clean_number(shares)
        clean_weight = clean_number(weight)
        
        holdings.append({
            "stock_code": code_str,
            "stock_name": name_str,
            "shares": int(clean_shares),
            "weight": float(clean_weight)
        })
        
    print(f"成功解析到 {len(holdings)} 檔成分股持股。")
    
    save_etf_data(etf_code, trading_date, summary_data, holdings)
    print(f"已成功將 {etf_code} 在 {trading_date} 的投組資料匯入資料庫。")
    return trading_date
